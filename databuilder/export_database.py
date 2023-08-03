import os
import sys
import json
import openpyxl

import _config
import _utils


def parse_constants(_worksheet, _data):
    _data[_worksheet.title] = list()
    count = _worksheet.cell(1, 3).value

    for row in _worksheet.iter_rows(min_row = 3, max_row = 3 + int(count - 1), min_col = 3, max_col = 5):
        row = [cell.value for cell in row]           
        
        item = {}
        item['Name'] = row[0]
        item['Value'] = row[1]
        
        _data[_worksheet.title].append(item)

static_db_sheet_func = [
    parse_constants
]

def export_staticdb():
    source = os.path.join(_config.DATA_DATABASE_PATH, 'Database.xlsx')
    output = os.path.join(_config.DATA_DATABASE_PATH, 'Database.json')
    
    book = openpyxl.load_workbook(source, data_only = True)   
    data = {}
    
    for sheet in book.worksheets:
        static_db_sheet_func[book.index(sheet)](sheet, data)
        
    with open(output, 'w') as f:
        json.dump(data, f)

def run():
    export_staticdb()